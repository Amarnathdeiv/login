import openpyxl

# excel_loc=r"C:\Users\mohan\Documents\amar\amar.xlsx"
# workbook=openpyxl.load_workbook(excel_loc)
# sheet=workbook.active
# # cellval=sheet.cell(2,1)
# # val=cellval.value
# # print(val)
# # C:\Users\mohan\PycharmProjects\amarselProject\Framebaseclass\amar.xlsx
#
#
#
# rows=sheet.max_row
# col=sheet.max_column
# print(rows)
# print(col)
#
# for i in range(1,rows+1):
#     for j in range(1,col+1):
#         cel=sheet.cell(i,j)
#         val=cel.value
#         print(val)

excel_location = "framework.xlsx"
book = openpyxl.Workbook()
sheet = book.active
cell = sheet.cell(1, 1)
cell.value = 'python'
book.save(excel_location)
print("done....")
